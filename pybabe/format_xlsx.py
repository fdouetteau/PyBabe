
from base import StreamHeader, BabeBase, StreamFooter


def valuenormalize(cell):
    "Build the row value out of a cell"
    if cell.number_format == '0':
        try:
            return int(cell.internal_value)
        except:
            return cell.internal_value
    else:
        return cell.internal_value


def read(format, stream,  kwargs):
    from openpyxl import load_workbook
    wb = load_workbook(filename=stream, use_iterators=True)
    ws = wb.get_active_sheet()
    it = ws.iter_rows()
    fields = kwargs.get('fields', None)
    if not fields:
        fields = [cell.internal_value for cell in it.next()]
    metainfo = StreamHeader(**dict(kwargs, fields=fields))
    yield metainfo
     # it brings a new method: iter_rows()
    for row in it:
        ## stop at the first row with "none"
        nrow = map(valuenormalize, row)
        if not any(nrow):
            break
        yield metainfo.t._make(nrow)
    yield StreamFooter()


def write(format, metainfo, instream, outfile, encoding, **kwargs):
    from openpyxl import Workbook
    wb = Workbook(optimized_write=True)
    ws = wb.create_sheet()
    ws.append(metainfo.fields)
    for k in instream:
        if isinstance(k, StreamFooter):
            break
        else:
            ws.append(list(k))
    wb.save(outfile)

BabeBase.addPullPlugin('xlsx', ['xlsx'], read, need_seek=True)
BabeBase.addPushPlugin('xlsx', ['xlsx'], write)
